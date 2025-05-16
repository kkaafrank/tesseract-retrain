from enum import Enum
from os import makedirs
from pathlib import Path
from re import IGNORECASE, compile, match

import openpyxl

from src import constants
from src.env.env import ENV

# string contains a japanese character
_JAPANESE_CHAR_REGEX = compile(
    #  hiragana     katakana     kanji        rare kanji
    r".*[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF\u3400-\u4DBF].*",
    flags=IGNORECASE,
)
_JAPANESE_SPACE_UNICODE_CHAR = "\u3000"

OUTPUT_FILE_BASE = ENV["MODEL_NAME"].replace("_", ".") + ".exp{}"
_OUTPUT_TEXT_NAME_FORMAT = OUTPUT_FILE_BASE + ".gt.txt"
OUTPUT_TEXT_NAME_GLOB = _OUTPUT_TEXT_NAME_FORMAT.format("*")


class InputTypes(Enum):
    """Valid input type enumeration"""

    EXCEL = "excel"
    TEXT = "text"

    @staticmethod
    def get_input_options() -> list[str]:
        return [
            InputTypes.EXCEL.value,
            InputTypes.TEXT.value,
        ]


# TODO: openpyxl type hints
def should_write_cell(cell_value) -> bool:
    """Whether or not the cell should be written to a text file.

    Args:
        cell_value: the value of the cell from openpyxl

    Returns:
        bool: if the cell should be written to text.
            Returns false if the value is None, is not an instance of str,
            or if there are no Japanese characters
    """
    if cell_value is None:
        return False

    if not isinstance(cell_value, str):
        return False

    if match(_JAPANESE_CHAR_REGEX, cell_value) is None:
        return False

    return True


def strip_spaces(input_string: str) -> str:
    """Removes space characters (normal and Japanese) from string

    Args:
        cell_text: string to clean

    Returns:
        str: text without Japanese space characters
    """
    text_to_write = input_string.replace(_JAPANESE_SPACE_UNICODE_CHAR, "")
    text_to_write = input_string.replace(" ", "")
    return text_to_write


def extract_japanese_text(input_file_path: Path, input_type: InputTypes) -> Path:
    """Extracts Japanese text from an excel or text file and writes each line to
    its own text file.

    Args:
        input_file_path: path to the input file to read
        input_type: input type. currently only excel is supported

    Returns:
        Path: folder containing all created text files
    """
    output_folder_name = constants.MODEL_NAME + "-ground-truth"
    output_folder_path = (
        Path(__file__).parent.parent / "tesstrain" / "data" / output_folder_name
    )
    makedirs(output_folder_path, exist_ok=True)

    if input_type == InputTypes.EXCEL:
        parse_excel_input(input_file_path, output_folder_path)
    # TODO: parsing from text file
    else:
        print("Input type not supported")

    return output_folder_path


def parse_excel_input(input_file_path: Path, output_folder_path: Path) -> None:
    """Parses and excel file for Japanese text and writes each line to a file

    Args:
        input_file_path: path to the excel file to be used as input
        output_folder_path: folder to write output files to
    """
    workbook = openpyxl.open(input_file_path)
    worksheet = workbook["Anki"]
    previous_line = ""
    num_files = 0
    max_files_reached = False

    for row_index in range(1, worksheet.max_row + 1, 1):
        for col_index in range(1, worksheet.max_column + 1, 1):
            cell_value = worksheet.cell(row_index, col_index).value
            if not should_write_cell(cell_value):
                continue

            text_to_write = cell_value
            if ENV["SHOULD_STRIP_JPN_SPACES"]:
                text_to_write = strip_spaces(cell_value)

            if text_to_write and previous_line != text_to_write:
                write_line_to_output_file(text_to_write, output_folder_path, num_files)

                previous_line = text_to_write
                num_files += 1

            if (
                ENV["LIMIT_INPUT_TEXT_FILES"]
                and num_files >= ENV["MAX_LIMITED_TEXT_FILES"]
            ):
                max_files_reached = True
                break

        if max_files_reached:
            break

        if row_index % 100 == 0:
            print(f"Finished extracting {row_index} rows of text from excel")


def write_line_to_output_file(
    line_to_write: str, parent_folder: Path, file_index: int
) -> None:
    """Writes a string to a text file

    Args:
        line_to_write (str): string to write
        parent_folder (Path): folder to write text file to
        file_index (int): index of the file to write. used to unique-ify each file name
    """
    next_output_file_name = _OUTPUT_TEXT_NAME_FORMAT.format(file_index)
    next_output_file_path = parent_folder / next_output_file_name
    with open(next_output_file_path, "w", encoding="utf-8") as output_file:
        output_file.write(line_to_write)
